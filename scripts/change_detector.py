# ==============================================================================
# change_detector.py — Detect URL changes between two JSON collections
# ==============================================================================
# Purpose: Compare two URL collections to identify new and deleted URLs
# Sections: Imports, Helpers, Main Logic, Public API

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")
    print("Install with: pip install openpyxl")


# ==============================================================================
# Helpers
# ==============================================================================

def resolve_file_path(filepath: str) -> str:
    """Resolve file path, handling external storage placeholders."""
    path = Path(filepath)
    
    # Check if this is a placeholder file (moved to external storage)
    if path.exists() and path.stat().st_size < 1000:  # Small file, might be placeholder
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if 'external_location' in data:
                    external_path = data['external_location']
                    if Path(external_path).exists():
                        print(f"Using external file: {external_path}")
                        return external_path
        except (json.JSONDecodeError, KeyError):
            pass
    
    return str(path)


def extract_site_id_from_path(filepath: str) -> str:
    """Extract site ID from filepath (e.g., 'judiciary_uk' from 'output/judiciary_uk_20250811_191111/full_url_set.json')."""
    path_parts = Path(filepath).parts
    for part in path_parts:
        if '_' in part and not part.startswith('20'):  # Skip timestamp parts
            return part
    raise ValueError(f"Could not extract site_id from path: {filepath}")


def extract_timestamp_from_path(filepath: str) -> str:
    """Extract timestamp from filepath (e.g., '20250811_203648' from 'output/judiciary_uk_20250811_203648/full_url_set.json')."""
    path_parts = Path(filepath).parts
    
    for part in path_parts:
        # Look for timestamp pattern: YYYYMMDD_HHMMSS (8 digits + underscore + 6 digits)
        if len(part) == 15 and part.startswith('20') and '_' in part:
            # Verify it's a valid timestamp format
            try:
                date_part, time_part = part.split('_')
                if len(date_part) == 8 and len(time_part) == 6:
                    # Validate that both parts are numeric
                    int(date_part)
                    int(time_part)
                    return part
            except (ValueError, AttributeError):
                continue
        
        # Handle embedded timestamp in directory names like 'judiciary_uk_20250811_203648'
        if '_' in part and len(part) > 15:
            # Look for timestamp pattern within the part
            for i in range(len(part) - 14):  # Need at least 15 chars for timestamp
                potential_timestamp = part[i:i+15]
                if (potential_timestamp.startswith('20') and 
                    '_' in potential_timestamp and 
                    len(potential_timestamp) == 15):
                    try:
                        date_part, time_part = potential_timestamp.split('_')
                        if len(date_part) == 8 and len(time_part) == 6:
                            # Validate that both parts are numeric
                            int(date_part)
                            int(time_part)
                            return potential_timestamp
                    except (ValueError, AttributeError):
                        continue
    
    raise ValueError(f"Could not extract timestamp from path: {filepath}")


def load_url_collection(filepath: str) -> Dict[str, Any]:
    """Load URL collection from JSON file."""
    # Resolve external file paths if needed
    resolved_path = resolve_file_path(filepath)
    
    try:
        with open(resolved_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found: {resolved_path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in file {resolved_path}: {e}")


def get_urls_from_collection(data: Dict[str, Any]) -> Dict[str, List[str]]:
    """Extract URLs and their detection methods from collection data."""
    url_info = {}
    
    if isinstance(data, dict) and 'urls' in data and isinstance(data['urls'], list):
        for item in data['urls']:
            if isinstance(item, dict) and 'url' in item:
                url = item['url']
                # Extract detection methods if available, default to empty list
                detection_methods = item.get('detection_methods', [])
                url_info[url] = detection_methods
    
    return url_info


def detect_changes(old_urls_info: Dict[str, List[str]], new_urls_info: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    """Detect new and deleted URLs between two collections, preserving detection methods."""
    changes = []
    
    # New URLs (in newer file, not in older)
    for url in new_urls_info.keys() - old_urls_info.keys():
        changes.append({
            "url": url, 
            "status": "new",
            "detection_methods": new_urls_info[url]
        })
    
    # Deleted URLs (in older file, not in newer)
    for url in old_urls_info.keys() - new_urls_info.keys():
        changes.append({
            "url": url, 
            "status": "removed",
            "detection_methods": old_urls_info[url]  # Important: preserve from older file
        })
    
    return changes


def save_changes_output(changes_data: Dict[str, Any], output_dir: str = "changes") -> str:
    """Save changes to timestamped JSON file in changes directory."""
    # Ensure output directory exists
    Path(output_dir).mkdir(exist_ok=True)
    
    # Generate filename: site_id_timestamp.json
    site_id = changes_data["site_id"]
    timestamp = changes_data["timestamp"]
    filename = f"{site_id}_{timestamp}.json"
    output_path = Path(output_dir) / filename
    
    # Save to file
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(changes_data, f, indent=2, ensure_ascii=False)
    
    return str(output_path)


def save_changes_excel(changes_data: Dict[str, Any], output_dir: str = "changes") -> str:
    """Save changes to Excel file with site_id as sheet name and timestamp in top row."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    # Ensure output directory exists
    Path(output_dir).mkdir(exist_ok=True)
    
    # Generate filename: site_id_timestamp.xlsx
    site_id = changes_data["site_id"]
    timestamp = changes_data["timestamp"]
    filename = f"{site_id}_{timestamp}.xlsx"
    output_path = Path(output_dir) / filename
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = site_id
    
    # Set timestamp in top row (A1)
    ws['A1'] = f"Timestamp: {timestamp}"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='left')
    
    # Set headers in row 3 (A3, B3, C3)
    ws['A3'] = "status"
    ws['B3'] = "url"
    ws['C3'] = "detection_methods"
    
    # Style headers with light blue background and borders
    from openpyxl.styles import PatternFill, Border, Side
    
    # Define colors and borders
    light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    light_grey_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    darker_grey_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    
    # Apply header styling
    for cell in ['A3', 'B3', 'C3']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].fill = light_blue_fill
        ws[cell].border = thick_bottom_border  # Thick border under header
    
    # Add data rows starting from row 4
    for i, change in enumerate(changes_data["urls"], start=4):
        # Apply alternating row colors
        row_fill = light_grey_fill if i % 2 == 0 else darker_grey_fill
        
        # Status column
        ws[f'A{i}'] = change["status"]
        ws[f'A{i}'].fill = row_fill
        ws[f'A{i}'].border = thin_border
        ws[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # URL column
        ws[f'B{i}'] = change["url"]
        ws[f'B{i}'].fill = row_fill
        ws[f'B{i}'].border = thin_border
        ws[f'B{i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Detection methods column
        detection_methods = change.get("detection_methods", [])
        if isinstance(detection_methods, list):
            # Join multiple methods with commas, or show as single string
            methods_text = ", ".join(detection_methods) if detection_methods else "N/A"
        else:
            methods_text = str(detection_methods) if detection_methods else "N/A"
        
        ws[f'C{i}'] = methods_text
        ws[f'C{i}'].fill = row_fill
        ws[f'C{i}'].border = thin_border
        ws[f'C{i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Style status column with colors
        if change["status"] == "new":
            ws[f'A{i}'].font = Font(color="008000", bold=True)  # Green for new
        elif change["status"] == "removed":
            ws[f'A{i}'].font = Font(color="FF0000", bold=True)  # Red for removed
    
    # Auto-adjust column widths to fit content
    # Status column - fixed width for status
    ws.column_dimensions['A'].width = 12
    
    # URL column - calculate optimal width based on content
    max_url_length = 0
    for change in changes_data["urls"]:
        url_length = len(change["url"])
        # Count each character as roughly 1.1 units (accounting for font)
        adjusted_length = url_length * 1.1
        max_url_length = max(max_url_length, adjusted_length)
    
    # Set URL column width (min 60, max 100 characters to make room for detection_methods)
    url_width = min(max(max_url_length, 60), 100)
    ws.column_dimensions['B'].width = url_width
    
    # Detection methods column - calculate optimal width
    max_methods_length = 0
    for change in changes_data["urls"]:
        detection_methods = change.get("detection_methods", [])
        if isinstance(detection_methods, list):
            methods_text = ", ".join(detection_methods) if detection_methods else "N/A"
        else:
            methods_text = str(detection_methods) if detection_methods else "N/A"
        
        methods_length = len(methods_text)
        adjusted_length = methods_length * 1.1
        max_methods_length = max(max_methods_length, adjusted_length)
    
    # Set detection methods column width (min 20, max 40 characters)
    methods_width = min(max(max_methods_length, 20), 40)
    ws.column_dimensions['C'].width = methods_width
    
    # Set row heights for better readability
    ws.row_dimensions[1].height = 25  # Timestamp row
    ws.row_dimensions[3].height = 20  # Header row
    for i in range(4, 4 + len(changes_data["urls"])):
        ws.row_dimensions[i].height = 18  # Data rows
    
    # Save workbook
    wb.save(output_path)
    wb.close()
    
    return str(output_path)


# ==============================================================================
# Main Logic
# ==============================================================================

def compare_url_collections(older_filepath: str, newer_filepath: str) -> Dict[str, Any]:
    """Compare two URL collections and return changes."""
    # Load both collections
    older_data = load_url_collection(older_filepath)
    newer_data = load_url_collection(newer_filepath)
    
    # Extract URLs from both collections
    older_urls_info = get_urls_from_collection(older_data)
    new_urls_info = get_urls_from_collection(newer_data)
    
    # Detect changes
    url_changes = detect_changes(older_urls_info, new_urls_info)
    
    # Use site_id from the newer file data, fallback to path extraction
    site_id = newer_data.get('site_id') or extract_site_id_from_path(newer_filepath)
    timestamp = extract_timestamp_from_path(newer_filepath)
    
    # Prepare output structure
    changes_data = {
        "site_id": site_id,
        "timestamp": timestamp,
        "urls": url_changes
    }
    
    return changes_data


# ==============================================================================
# Public API
# ==============================================================================

__all__ = [
    'compare_url_collections',
    'detect_changes',
    'save_changes_output',
    'save_changes_excel',
]


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) != 3:
        print("Usage: python change_detector.py <older_filepath> <newer_filepath>")
        print("Example: python change_detector.py output/judiciary_uk_20250810_190811/full_url_set.json output/judiciary_uk_20250811_191111/full_url_set.json")
        sys.exit(1)
    
    older_filepath = sys.argv[1]
    newer_filepath = sys.argv[2]
    
    try:
        # Compare collections
        changes = compare_url_collections(older_filepath, newer_filepath)
        
        # Save JSON output
        json_output_path = save_changes_output(changes)
        print(f"JSON output saved to: {json_output_path}")
        
        # Save Excel output
        try:
            excel_output_path = save_changes_excel(changes)
            print(f"Excel output saved to: {excel_output_path}")
        except ImportError as e:
            print(f"Excel export skipped: {e}")
        except Exception as e:
            print(f"Excel export failed: {e}")
        
        print(f"\nChanges detected: {len(changes['urls'])} URLs")
        print(f"New URLs: {len([u for u in changes['urls'] if u['status'] == 'new'])}")
        print(f"Removed URLs: {len([u for u in changes['urls'] if u['status'] == 'removed'])}")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
